import zipfile
from io import TextIOWrapper
import openpyxl
from test_homework.page_object import zip_file
from pypdf import PdfReader
import csv


def test_zip(test_zip_create):
    with zipfile.ZipFile(zip_file, "r") as zip_f:
        zip_f = len(zip_f.filelist)
    test_files = [zip_file]
    for file in test_files:
        assert "ZIP status for {0}: {1}".format(file, zipfile.is_zipfile(file))
        assert zip_f == 3
    with zipfile.ZipFile(zip_file) as zf:
        ok = zf.testzip()
        assert "It`s OK!" if ok is None else print(f"{ok} is fail.")


def test_pdf():
    with zipfile.ZipFile(zip_file, "r") as zip_f:
        with zip_f.open("file_pdf.pdf") as pdf_f:
            pdf_text = PdfReader(pdf_f)
            page = len(pdf_text.pages)
            assert page == 1
            assert "PDF" in pdf_text.pages[0].extract_text(), f"Это PDF файл"


def test_add_xslx():
    with zipfile.ZipFile(zip_file, "r") as zip_f:
        with zip_f.open("file_xlsx.xlsx") as xsl_f:
            wb = openpyxl.load_workbook(xsl_f)
            sheet = wb.active
            cells = sheet["A1":"B6"]
            a1 = sheet.cell(row=1, column=1).value
            a2 = sheet.cell(row=2, column=1).value
            a3 = sheet.cell(row=3, column=1).value
            for c1, c2 in cells:
                assert "{0:8} {1:8}".format(c1.value, c2.value)
            assert a1 == "Название"
            assert a2 == "Филип Котлер. Основы маркетинга"
            assert a3 != "Йона Бергер. Заразительный"


def test_add_csv():
    with zipfile.ZipFile(zip_file, "r") as zip_f:
        with zip_f.open("file_csv.csv") as csv_f:
            csv_text = list(csv.reader(TextIOWrapper(csv_f, "utf-8-sig")))
            assert "Название" == csv_text[0][0]
            assert "Игорь Манн. Маркетинг на 100%" == csv_text[12][0]
